from openpyxl import Workbook
from django.http import Http404, HttpResponse
from django.shortcuts import get_object_or_404
from payments.models import Consultation
from .models import ConsultantSchedule

def export_form(request, id, form_attr, filename):

    schedule = get_object_or_404(
        ConsultantSchedule,
        id=id,
        consultant=request.user.user_consultant,
    )

    consultation = get_object_or_404(
        Consultation.objects.select_related(
            "service__student"
        ),
        schedule=schedule,
    )

    student = consultation.service.student

    if not hasattr(student, form_attr):
        raise Http404("فرم مورد نظر تکمیل نشده است.")

    form = getattr(student, form_attr)

    wb = Workbook()
    ws = wb.active
    ws.title = filename.split(".")[0]

    # عنوان ستون‌ها
    ws["A1"] = "عنوان"
    ws["B1"] = "مقدار"

    row = 2

    for field in form._meta.fields:

        if field.name in ["id", "student"]:
            continue

        value = getattr(form, field.name)

        # حذف * از verbose_name
        label = field.verbose_name.replace("*", "")

        # اگر فایل بود فقط لینک/نام فایل ثبت شود
        if hasattr(value, "url"):
            value = request.build_absolute_uri(value.url)

        ws.cell(row=row, column=1).value = label
        ws.cell(row=row, column=2).value = str(value) if value is not None else ""

        row += 1

    # تنظیم عرض ستون‌ها
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 80

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response[
        "Content-Disposition"
    ] = f'attachment; filename="{filename}"'

    wb.save(response)

    return response

def export_form1_excel(request, id):
    return export_form(
        request,
        id,
        "student_form_1",
        "student_form_1.xlsx",
    )


def export_form2_excel(request, id):
    return export_form(
        request,
        id,
        "student_form_2",
        "student_form_2.xlsx",
    )


def export_form3_excel(request, id):
    return export_form(
        request,
        id,
        "student_form_3",
        "student_form_3.xlsx",
    )